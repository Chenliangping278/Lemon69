# -coding: UTF-8 -*-
# @Time  : 2020/06/24 20:01
# @Author: Liangping_Chen
# @E-mail: chenliangping_2018@foxmail.com


from openpyxl import load_workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]#储存所有行的测试用例数据
    for i in range(2,sheet.max_row+1):
        case=[]#某一行测试用例数据
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i,column=j).value)
        all_case.append(case)
    print(sheet.max_row)
    return all_case#返回所有测试用例数据

    #print(all_case)
read_data('test_case.xlsx', 'recharge')






def write_data(file_name,sheet_name,row,column,value):#此函数为写入结果到Excel中
    #开始写入结果
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    #定位单元格存值 行，列，值
    sheet.cell(row=row,column=column).value=value
    #进行判断，期望值与实际值是否相等，判断用例是否执行通过
    wb.save('test_case.xlsx')
